"""
Vercel serverless function — NFS-e download endpoint.
Returns a ZIP containing all NFS-e XMLs + an XLSX summary.
Certificate is never persisted — only used during the request lifecycle.
"""

import sys, os, io, json, zipfile, tempfile, re
import xml.etree.ElementTree as ET
from datetime import date

from flask import Flask, request, send_file, jsonify

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from app.services.cert_handler import export_to_pem_files, get_cert_info
from app.services.nfse_nacional import NfseNacionalClient
from app.services import cert_handler as _ch
from app.services.auth_service import verify_token
from app.services.subscription_service import verificar_e_registrar_download
from app.services.db import execute
from app.services.portal_reconciliation import (
    listar_chaves_recebidas_portal, reconciliar,
)
from app.services.cnpj_lookup import lookup_cnpj

app = Flask(__name__)
# Limite do corpo total da requisição (certificado ~1 MB + campos do formulário).
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024  # 2 MB

MAX_CERT_BYTES = 1 * 1024 * 1024  # 1 MB


def _log_download(user_id: str, cnpj: str, sucesso: bool, erro: str = None) -> None:
    try:
        execute(
            "INSERT INTO download_logs (user_id, cnpj, sucesso, erro) VALUES (%s, %s, %s, %s)",
            (user_id, cnpj, sucesso, erro),
        )
    except Exception:
        pass


# ── NFS-e XML parser ─────────────────────────────────────────────────────────

NS = "http://www.sped.fazenda.gov.br/nfse"

_digits = lambda s: "".join(c for c in s if c.isdigit())


def _txt(root, *tags) -> str:
    """Find first matching tag (with or without namespace) and return its text."""
    for tag in tags:
        for candidate in (f"{{{NS}}}{tag}", tag):
            el = root.find(f".//{candidate}")
            if el is not None and el.text:
                return el.text.strip()
    return ""


def parse_nfse(chave: str, nsu: int, xml: str, cnpj_contribuinte: str = "",
               is_cancelada: bool = False) -> dict:
    row = {
        "ChaveAcesso": chave,
        "NSU": nsu,
        "Tipo": "Indefinida",   # Emitida / Recebida / Indefinida
        "Cancelada": "Sim" if is_cancelada else "Não",
        "NumeroNFSe": "",
        "DataEmissao": "",
        "Competencia": "",
        "CNPJPrestador": "",
        "NomePrestador": "",
        "CNPJTomador": "",
        "NomeTomador": "",
        "DescricaoServico": "",
        "ValorServico": "",
        "ValorISSQN": "",
        "Aliquota": "",
        "MunicipioIncidencia": "",
    }
    try:
        root = ET.fromstring(xml)
        row["NumeroNFSe"]          = _txt(root, "nNFSe")
        row["DataEmissao"]         = _txt(root, "dhEmi", "dEmi", "DataEmissao")[:10]
        _comp_raw = _txt(root, "dComp", "dtCompetencia", "DataCompetencia",
                         "competencia", "Competencia", "CompNfse",
                         "dCompetencia", "dhCompetencia")[:10]
        # Regex fallback: find dComp tag regardless of XML namespace prefix
        if not _comp_raw:
            m = re.search(r'<(?:[^:>\s]+:)?dComp>(\d{4}-\d{2}(?:-\d{2})?)<', xml)
            if m:
                _comp_raw = m.group(1)[:10]
        # Final fallback: use emission month when dComp absent from XML
        if not _comp_raw and row["DataEmissao"]:
            _comp_raw = row["DataEmissao"][:7]
        # Format as MM/AAAA (matches portal display; sortable and filterable in Excel)
        if _comp_raw:
            _parts = _comp_raw.split("-")
            if len(_parts) >= 2 and len(_parts[0]) == 4:
                row["Competencia"] = f"{_parts[1]}/{_parts[0]}"  # YYYY-MM → MM/YYYY
            else:
                row["Competencia"] = _comp_raw
        row["CNPJPrestador"]       = _txt(root, "emit//CNPJ", "prestador//CNPJ",
                                          "prest//CNPJ", "Prestador//CNPJ")
        row["NomePrestador"]       = _txt(root, "emit//xNome", "prestador//xNome",
                                          "prest//xNome", "xNome")
        row["CNPJTomador"]         = _txt(root, "tomador//CNPJ", "dest//CNPJ",
                                          "Tomador//CNPJ")
        row["NomeTomador"]         = _txt(root, "tomador//xNome", "dest//xNome")
        row["DescricaoServico"]    = _txt(root, "xDescServ", "descricao", "Descricao")
        row["ValorServico"]        = _txt(root, "vServ", "vServPrest//vReceb", "ValorServicos")
        row["ValorISSQN"]          = _txt(root, "vISSQN", "vIss", "ValorIss")
        row["Aliquota"]            = _txt(root, "pAliq", "aliquota", "Aliquota")
        row["MunicipioIncidencia"] = _txt(root, "cLocIncid", "cMunFG", "CodigoMunicipio")

        # Fallback: CNPJ do prestador pode estar no topo do XML
        if not row["CNPJPrestador"]:
            row["CNPJPrestador"] = _txt(root, "CNPJ")

        # Detect cancellation — two layers:
        # Layer 1: TipoDocumento already set via is_cancelada before entering here.
        # Layer 2: raw text scan of XML string (namespace-agnostic, covers all variants).
        if row["Cancelada"] == "Não":
            xl = xml.lower()
            if any(p in xl for p in (
                "<dhcanc", "<infcancelamento", "<infcanc>", "<infcanc ",
                "<pedcanc", "<dtcancelamento", "<infcancnfse",
                "<confirmacaocancelamento", ">cancelada<", ">cancelado<",
            )):
                row["Cancelada"] = "Sim"
    except ET.ParseError:
        pass

    # ── Classify: Emitida / Recebida / Indefinida ─────────────────────────
    # Rule: based solely on CNPJPrestador vs the client's CNPJ.
    #   Emitida  → client IS the prestador (they issued the invoice)
    #   Recebida → another CNPJ is the prestador (issued BY someone else TO/for client)
    #   Indefinida → can't extract prestador CNPJ from the XML
    if cnpj_contribuinte:
        cnpj_d  = _digits(cnpj_contribuinte)
        prest_d = _digits(row["CNPJPrestador"])
        tom_d   = _digits(row["CNPJTomador"])

        if prest_d == cnpj_d:
            row["Tipo"] = "Emitida"          # client emitted this note
        elif prest_d:
            row["Tipo"] = "Recebida"         # another prestador emitted it
        elif tom_d == cnpj_d:
            row["Tipo"] = "Recebida"         # client is tomador → received
        # else: prestador unknown, tomador unknown → stays "Indefinida"

    return row


def _filter_by_tipo(results: list, xlsx_rows: list, tipo: str) -> tuple:
    """
    Filter results and xlsx_rows by NFS-e type:
    • "todas"    → keep everything
    • "emitidas" → only Emitida  (strict: client must be prestador)
    • "recebidas"→ Recebida + Indefinida (conservative: keep unknowns to avoid losses)
    """
    if tipo == "todas":
        return results, xlsx_rows

    keep = {
        "emitidas":  {"Emitida", "Indefinida"},   # conservative: don't discard unclassifiable
        "recebidas": {"Recebida", "Indefinida"},
    }.get(tipo, {"Emitida", "Recebida", "Indefinida"})

    paired = [(r, row) for r, row in zip(results, xlsx_rows) if row["Tipo"] in keep]
    if not paired:
        return [], []
    res, rows = zip(*paired)
    return list(res), list(rows)


# ── XLSX builder ─────────────────────────────────────────────────────────────

COLUNAS = [
    ("ChaveAcesso",        "Chave de Acesso",          50),
    ("NSU",                "NSU",                       8),
    ("Tipo",               "Tipo",                     12),  # Emitida / Recebida / Indefinida
    ("Cancelada",          "Cancelada",                10),  # Sim / Não
    ("NumeroNFSe",         "Número NFS-e",             14),
    ("DataEmissao",        "Data Emissão",             14),
    ("Competencia",        "Competência",              14),
    ("CNPJPrestador",      "CNPJ Prestador",           20),
    ("NomePrestador",      "Nome Prestador",           40),
    ("CNPJTomador",        "CNPJ Tomador",             20),
    ("NomeTomador",        "Nome Tomador",             40),
    ("DescricaoServico",   "Descrição Serviço",        40),
    ("ValorServico",       "Valor Serviço (R$)",       18),
    ("ValorISSQN",         "Valor ISSQN (R$)",         18),
    ("Aliquota",           "Alíquota (%)",             14),
    ("MunicipioIncidencia","Município Incidência",     20),
]


def build_xlsx(rows: list, faltantes: list | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F6AA5")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="EBF2FB")

    # Header
    for col, (_, label, width) in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Fills for "Tipo" column
    tipo_fills = {
        "Emitida":   PatternFill("solid", fgColor="D4EDDA"),  # light green
        "Recebida":  PatternFill("solid", fgColor="D1ECF1"),  # light blue
        "Indefinida":PatternFill("solid", fgColor="FFF3CD"),  # light yellow
    }
    cancelada_fill = PatternFill("solid", fgColor="F8D7DA")   # light red for cancelled

    # Data rows
    for r, row in enumerate(rows, 2):
        alt = r % 2 == 0
        is_row_cancelada = row.get("Cancelada") == "Sim"
        for col, (key, _, _) in enumerate(COLUNAS, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if key == "Tipo":
                cell.fill = tipo_fills.get(val, alt_fill if alt else PatternFill())
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "Cancelada":
                if is_row_cancelada:
                    cell.fill = cancelada_fill
                    cell.font = Font(bold=True, color="721C24", size=10)
                else:
                    cell.fill = alt_fill if alt else PatternFill()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif alt:
                cell.fill = alt_fill

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Aba extra: notas recebidas que constam no portal mas não vieram (reconciliação)
    if faltantes:
        add_sheet_faltantes(wb, faltantes)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Relatório de notas faltantes (reconciliação com o portal) ─────────────────

def _fmt_cnpj(cnpj: str) -> str:
    c = "".join(ch for ch in str(cnpj) if ch.isdigit())
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"
    return cnpj


COLUNAS_FALTANTES = [
    ("chave",          "Chave de Acesso",        55),
    ("cnpj_prestador", "CNPJ Prestador",         20),
    ("nome_prestador", "Nome Prestador",         40),
    ("municipio",      "Município Emissor (IBGE)", 22),
    ("como_obter",     "Como Obter",             60),
]

OBS_FALTANTE = (
    "Consta no portal nacional, mas não foi distribuída ao seu certificado (DF-e). "
    "Baixe manualmente em www.nfse.gov.br (Notas Recebidas) ou solicite o XML ao prestador."
)


def add_sheet_faltantes(wb, faltantes: list) -> None:
    """Add a 'Notas Faltantes' worksheet to the workbook, styled like the main
    sheet. Chave de Acesso is written as TEXT (Excel would mangle a 50-digit
    number into scientific notation, losing digits)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Notas Faltantes")

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="B45309")   # âmbar — sinaliza atenção
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="FEF3C7")       # âmbar claro (zebra)

    for col, (_, label, width) in enumerate(COLUNAS_FALTANTES, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment, cell.border = header_align, thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for r, f in enumerate(faltantes, 2):
        alt = r % 2 == 0
        valores = {
            "chave":          str(f.get("chave", "")),
            "cnpj_prestador": _fmt_cnpj(f.get("cnpj_prestador", "")),
            "nome_prestador": f.get("nome_prestador", "") or "(consulte pelo CNPJ)",
            "municipio":      f.get("municipio", ""),
            "como_obter":     OBS_FALTANTE,
        }
        for col, (key, _, _) in enumerate(COLUNAS_FALTANTES, 1):
            cell = ws.cell(row=r, column=col, value=valores[key])
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(key == "como_obter"))
            if key in ("chave", "municipio"):
                cell.number_format = "@"   # força TEXTO — preserva os 50 dígitos
            if alt:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions


# ── Main endpoint ─────────────────────────────────────────────────────────────

@app.route("/api/download", methods=["POST"])
def download():
    # ── Common fields ──────────────────────────────────────────────────────────
    auth_method = request.form.get("auth_method", "cert")   # "cert" | "login" (login temporariamente desativado)
    if auth_method == "login":
        return jsonify({"error": "O acesso por login e senha do portal está temporariamente desativado. "
                                  "Por enquanto, utilize o certificado digital A1 (.pfx/.p12)."}), 400
    cnpj        = "".join(c for c in request.form.get("cnpj", "") if c.isdigit())
    nome        = request.form.get("nome", "Cliente")
    data_ini_s  = request.form.get("data_inicial", "")
    data_fim_s  = request.form.get("data_final", "")
    ambiente    = request.form.get("ambiente", "producao")
    nsu_ini     = int(request.form.get("nsu_inicial", 0) or 0)
    tipo_nfse   = request.form.get("tipo_nfse", "todas")
    if tipo_nfse not in ("todas", "emitidas", "recebidas"):
        tipo_nfse = "todas"

    if len(cnpj) != 14:
        return jsonify({"error": "CNPJ inválido. Informe os 14 dígitos."}), 400

    # ── Auth check ──────────────────────────────────────────────────────────────
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return jsonify({"error": "Autenticação necessária."}), 401
    jwt_payload = verify_token(auth_header[7:])
    if not jwt_payload:
        return jsonify({"error": "Token inválido ou expirado. Faça login novamente."}), 401

    user_id = jwt_payload["sub"]

    # ── CNPJ must be a registered client (or already used before this rule) ──
    # Enforced server-side regardless of what the UI allows — closes the
    # loophole of posting an arbitrary CNPJ directly to this endpoint.
    # Grandfather clause: any CNPJ with prior successful-usage history for this
    # user is still allowed, so existing active users are never locked out by
    # this change even if they never registered a client before.
    known_cnpj = execute(
        "SELECT 1 FROM clients WHERE user_id = %s AND cnpj = %s "
        "UNION SELECT 1 FROM monthly_usage WHERE user_id = %s AND cnpj = %s "
        "LIMIT 1",
        (user_id, cnpj, user_id, cnpj),
        fetch="one",
    )
    if not known_cnpj:
        return jsonify({
            "error": "Este CNPJ ainda não está cadastrado. Cadastre-o na aba "
                     "\"Clientes\" antes de baixar as NFS-e.",
        }), 400

    # ── Subscription / trial check ──────────────────────────────────────────────
    allowed, msg = verificar_e_registrar_download(user_id, cnpj)
    if not allowed:
        return jsonify({"error": msg}), 403

    try:
        data_ini = date.fromisoformat(data_ini_s)
        data_fim = date.fromisoformat(data_fim_s)
    except ValueError:
        return jsonify({"error": "Datas inválidas. Use AAAA-MM-DD."}), 400

    tmp_cert_path = None
    messages = []

    try:
        # ── Authentication (Certificado Digital A1 — único método disponível) ──
        cert_file = request.files.get("cert")
        if not cert_file:
            return jsonify({"error": "Arquivo de certificado (.pfx) não enviado."}), 400

        password = request.form.get("password", "")

        # ── File size guard ─────────────────────────────────────────────
        cert_bytes = cert_file.read()
        if len(cert_bytes) > MAX_CERT_BYTES:
            return jsonify({"error": "Certificado muito grande (máx. 1 MB)."}), 400

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pfx", prefix="nfse_up_")
        tmp.write(cert_bytes)
        tmp.close()
        tmp_cert_path = tmp.name

        try:
            info = get_cert_info(tmp_cert_path, password)
            if info.get("vencido"):
                _log_download(user_id, cnpj, False, "Certificado vencido")
                return jsonify({"error": f"Certificado vencido em {info['validade']}."}), 400
        except Exception:
            _log_download(user_id, cnpj, False, "Certificado inválido ou senha incorreta")
            return jsonify({"error": "Certificado inválido ou senha incorreta."}), 400

        client = NfseNacionalClient(
            cnpj=cnpj, cert_path=tmp_cert_path,
            cert_password=password, ambiente=ambiente,
        )

        # ── Download NFS-e ─────────────────────────────────────────────────────
        try:
            results = client.consultar_por_periodo(
                data_inicial=data_ini, data_final=data_fim,
                log=messages.append, nsu_inicial=nsu_ini,
                tipo_scraping=tipo_nfse,
            )
        except RuntimeError as e:
            _log_download(user_id, cnpj, False, str(e))
            return jsonify({"error": str(e), "log": messages}), 400

        if not results:
            _log_download(user_id, cnpj, True, "Nenhuma NFS-e encontrada")
            return jsonify({
                "error": "Nenhuma NFS-e encontrada para o período informado.",
                "log": messages,
            }), 404

        # ── Parse + classify every note ────────────────────────────────────
        xlsx_rows = []
        for i, entry in enumerate(results):
            chave        = entry[0]
            xml          = entry[1]
            is_cancelada = bool(entry[2]) if len(entry) > 2 else False
            nsu_val      = i + 1
            xlsx_rows.append(parse_nfse(chave, nsu_val, xml,
                                        cnpj_contribuinte=cnpj,
                                        is_cancelada=is_cancelada))

        # Conjunto de TODAS as chaves baixadas (antes do filtro de tipo), em 44
        # dígitos — base da reconciliação com o portal (evita falso-positivo).
        chaves_baixadas_all = {str(e[0])[:44] for e in results}

        # ── Apply tipo filter (after full download — never miss a note) ────
        total_baixadas = len(results)
        results, xlsx_rows = _filter_by_tipo(results, xlsx_rows, tipo_nfse)

        tipo_label = {"emitidas": "emitidas", "recebidas": "recebidas"}.get(tipo_nfse, "")
        if not results:
            _log_download(user_id, cnpj, True, "Filtro tipo removeu tudo")
            tipo_msg = f" {tipo_label}" if tipo_label else ""
            return jsonify({
                "error": (
                    f"Nenhuma NFS-e{tipo_msg} encontrada no período informado "
                    f"(total baixado: {total_baixadas}, nenhuma classificada como {tipo_label or 'qualquer tipo'})."
                ),
                "log": messages,
            }), 404

        # ── Summary counts by type ──────────────────────────────────────────
        contagem = {"Emitida": 0, "Recebida": 0, "Indefinida": 0, "Cancelada": 0}
        for row in xlsx_rows:
            contagem[row["Tipo"]] = contagem.get(row["Tipo"], 0) + 1
            if row.get("Cancelada") == "Sim":
                contagem["Cancelada"] += 1
        n_canceladas = contagem["Cancelada"]
        n_ativas     = len(xlsx_rows) - n_canceladas
        cancelada_info = f" | Canceladas (excluídas dos XMLs): {n_canceladas}" if n_canceladas else ""
        messages.append(
            f"  Resumo: {len(xlsx_rows)} nota(s) no Excel "
            f"[Emitidas: {contagem['Emitida']} | Recebidas: {contagem['Recebida']} | Indefinidas: {contagem['Indefinida']}{cancelada_info}]"
            + (f" — filtro: {tipo_label}" if tipo_label else "")
        )
        if n_canceladas:
            messages.append(
                f"  ZIP: {n_ativas} XML(s) incluído(s) — {n_canceladas} nota(s) cancelada(s) removida(s)."
            )

        # ── Reconciliação: notas recebidas que constam no portal mas não foram
        #    distribuídas via certificado/DF-e. Best-effort: NUNCA derruba o
        #    download — qualquer falha é registrada e o fluxo segue normal. ──
        notas_faltantes = []
        if tipo_nfse in ("todas", "recebidas"):
            try:
                chaves_portal = listar_chaves_recebidas_portal(
                    tmp_cert_path, password, data_ini, data_fim, log=messages.append,
                )
                notas_faltantes = reconciliar(
                    chaves_portal, chaves_baixadas_all,
                    enrich_nome=lambda c: (lookup_cnpj(c) or {}).get("nome", ""),
                )
                if notas_faltantes:
                    messages.append(
                        f"  ⚠ {len(notas_faltantes)} nota(s) recebida(s) constam no portal mas "
                        f"não vieram pelo certificado — ver aba 'Notas Faltantes' na planilha."
                    )
                else:
                    messages.append("  Reconciliação com o portal: nenhuma nota recebida faltante.")
            except Exception as e:
                messages.append(f"  Reconciliação indisponível (download não afetado): {e}")

        # ── Build ZIP (XMLs + XLSX) ─────────────────────────────────────────
        safe_nome   = "".join(c for c in nome if c.isalnum() or c in " _-")[:40]
        periodo     = f"{data_ini.strftime('%Y%m')}-{data_fim.strftime('%Y%m')}"
        tipo_suffix = {"emitidas": "_Emitidas", "recebidas": "_Recebidas"}.get(tipo_nfse, "")
        zip_name    = f"NFS-e_{cnpj}_{safe_nome}_{periodo}{tipo_suffix}.zip"

        zip_buf = io.BytesIO()
        seen: set = set()

        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # XML files — only active (non-cancelled) notes
            for entry, row in zip(results, xlsx_rows):
                if row.get("Cancelada") == "Sim":
                    continue
                chave = entry[0]
                xml   = entry[1]
                safe_chave = "".join(c for c in str(chave) if c.isalnum())[:44]
                base  = f"xml/NFS-e_{safe_chave}.xml"
                fname = base
                suf   = 1
                while fname in seen:
                    fname = f"xml/NFS-e_{safe_chave}_{suf}.xml"
                    suf  += 1
                seen.add(fname)
                content = xml.encode("utf-8") if isinstance(xml, str) else xml
                zf.writestr(fname, content)

            # XLSX (inclui a aba "Notas Faltantes" quando a reconciliação achou lacunas)
            xlsx_bytes = build_xlsx(xlsx_rows, faltantes=notas_faltantes)
            zf.writestr(f"NFS-e_{cnpj}_{periodo}{tipo_suffix}.xlsx", xlsx_bytes)

        zip_buf.seek(0)
        _log_download(user_id, cnpj, True)
        resp = send_file(zip_buf, mimetype="application/zip",
                         as_attachment=True, download_name=zip_name)
        # Header lido pelo frontend para exibir o aviso de notas faltantes.
        resp.headers["X-Notas-Faltantes"] = str(len(notas_faltantes))
        return resp

    finally:
        if tmp_cert_path:
            try:
                os.unlink(tmp_cert_path)
            except OSError:
                pass
        for f in list(_ch._temp_files):
            try:
                os.unlink(f)
                _ch._temp_files.remove(f)
            except (OSError, ValueError):
                pass
